import openpyxl
wb = openpyxl.load_workbook('input.xlsx')
sheet = wb.active
wb1 = openpyxl.Workbook()
wb1.create_sheet()
sheet1 = wb1.active
for row in range(sheet.max_row):
    for cell in range(sheet.max_column):
        x=sheet.cell(row+1,cell+1).value
        sheet1.cell(cell+1,row+1,x)
wb1.save('outpu.xlsx')
